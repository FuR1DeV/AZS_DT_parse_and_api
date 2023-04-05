import time
from random import randint
from datetime import datetime

import requests
from openpyxl import load_workbook
from xls2xlsx import XLS2XLSX
from rest_framework import viewsets

from database import db


class ParserViewSet(viewsets.ModelViewSet):

    @staticmethod
    def main_parser():
        """Главная функция парсинга данных, и обновление БД"""
        print("30 дней прошло. Пора обновлять данные")
        url = 'http://rss.tatneft.ru'
        response = requests.get(url)
        """Если статус код 200 значит ответ получен, и мы можем парсить сайт"""
        if response.status_code == 200:
            """Сохраняем список из url адресов"""
            urls = ["https://rss.tatneft.ru/prod", "https://rss.tatneft.ru/news", "https://rss.tatneft.ru/klient",
                    "https://rss.tatneft.ru/oil_cards", "https://rss.tatneft.ru/klient/materials",
                    "https://rss.tatneft.ru/infromation_provision", "https://rss.tatneft.ru/klient/chuzhie",
                    "https://rss.tatneft.ru/partners", "https://rss.tatneft.ru/locator",
                    "https://rss.tatneft.ru/locator/list"]
            """Имитируем поведение пользователя"""
            for index in range(len(urls)):
                requests.get(url=urls[index])
                time.sleep(randint(5, 11))
                print(f"click... {index + 1}")
            """Скачиваем нужный файл"""
            get_xml_list = requests.get(url="https://rss.tatneft.ru/outer/azs/get_xml_list").content
            print("Файл успешно скачен!")
            """Записываем его в файл с форматом xls"""
            with open(f"dt/xls_files/azs_monthly_prices_{datetime.now().month}.xls", "wb") as file:
                file.write(get_xml_list)
            """Конвертируем его в xlsx"""
            x2x = XLS2XLSX(f"dt/xls_files/azs_monthly_prices_{datetime.now().month}.xls")
            x2x.to_xlsx(f"dt/xls_files/azs_monthly_prices_{datetime.now().month}.xlsx")
            """Открываем новый файл xlsx"""
            book = load_workbook(filename=f"dt/xls_files/azs_monthly_prices_{datetime.now().month}.xlsx")
            sheet = book["Лист1"]
            """Перед тем как обновить таблицу БД,
            старые записи в таблице удаляем"""
            db.update_obj.delete()
            print("Обновляем Базы Данных")
            for row in range(2, sheet.max_row + 1):
                db.update_obj.update(latitude=sheet["E" + str(row)].value,
                                     longitude=sheet["F" + str(row)].value,
                                     dt=sheet["M" + str(row)].value,
                                     dt_taneko=sheet["N" + str(row)].value,
                                     dt_winter=sheet["V" + str(row)].value,
                                     dt_arctic=sheet["X" + str(row)].value)
        else:
            with open('error.txt', 'w') as f:
                f.write(response.content.decode("utf-8"))
